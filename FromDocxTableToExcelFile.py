
"""
This script is designed to process a docx file that contains multiple tables one each page.
The title of each page is extracted and served as identifier for the data on that page.
The contents of each table will be extracted and formatted into the desired format and exported into output excel files.
Please take a look of the example docx file to learn how this script work.
"""


import docx
import openpyxl
from openpyxl import Workbook
import re


"""
Since the auto numbering in the word file (table 3 on each page of example file) can not be recognized as text in python.
We need to convert the numbeirng to text first.
So, there is one step to do in the word file, before running this script.
Using VBA:
Open the document in Word.
Press Alt + F11 to open the VBA editor.
Press Ctrl + G to open the Immediate window.
Type ActiveDocument.ConvertNumbersToText and press Enter.
Close the VBA editor. The automatic numbering will be converted to regular text. 
"""


# This is designed to read the whole file and add them into the components, with label of Paragraph or Table.
# Later, the components will be read and processed. 
# Here, using XPath to access and assemble paragraphs and tables from scratch in the provided code allows for more granular and flexible control over the document structure. 
# Read more about using Xpath vs python-docx provided high-level objects such as 'Paragraph' and 'Table'
def list_docx_components(docx_path):
    doc = docx.Document(docx_path)
    components = []

    for element in doc.element.body:
        if element.tag.endswith('p'):
            para = element.xpath('.//w:t')
            para_text = ''.join([text.text for text in para]).strip()
            components.append(('Paragraph', para_text))
        elif element.tag.endswith('tbl'):
            table = []
            for row_idx, row in enumerate(element.xpath('.//w:tr')):
                row_data = []
                for col_idx, cell in enumerate(row.xpath('.//w:tc')):
                    cell_text = ''.join([text.text for text in cell.xpath('.//w:t')])
                    row_data.append(cell_text)
                table.append(row_data)
            components.append(('Table', table))

    return components

# This function deal with the first type of table, the table 1 on each page. 
# Write the pre defined headers and write the data to the file.
def write_location_details_to_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Location Details"
    
    headers = [
        "Location ID", "Project No.", "Project", "Client", "Date", "Location", "Logged By", 
        "Excavation Contractor", "Surface Elevation", "Excavation Equipment", "Completion Depth", 
        "Groundwater Depth First Encountered", "Groundwater Depth End of Excavation", 
        "Groundwater Elevation First Encountered", "Groundwater Elevation End of Excavation", 
        "Seasonal High Groundwater Depth", "SHGW Elevation"
    ]
    
    ws.append(headers)
    
    for entry in data:
        ws.append(entry)
    
    wb.save(output_path)

# This function deal with the second type of table, the table 2 on each page. 
# Write the pre defined headers and write the data to the file.
def write_field_geological_descriptions_to_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Geological Descriptions"
    
    headers = ["Location ID", "Depth Top", "Depth Base", "Description", "Legend Code"]
    ws.append(headers)
    
    for entry in data:
        ws.append(entry)
    
    wb.save(output_path)

# This function deal with the third type of table, the table 3 on each page. 
# Write the pre defined headers and write the data to the file.
def write_depth_related_remarks_to_excel(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Depth Related Remarks"
    
    headers = ["Location ID", "Depth Top", "Depth Base", "Remarks"]
    ws.append(headers)
    
    for entry in data:
        ws.append(entry)
    
    wb.save(output_path)


# Actually, consider editing the previous 3 functions into one write_to_excel can simplify this scripts

# This is for reformat the color description. Since the original document put color after the numbers, we need separte them and move the color forward.
# pay attention to the and/or situation. 
def reformat_color_description(description):
    # Split the description by 'and' or 'to' to handle multiple color parts
    parts = re.split(r'\s+and\s+|\s+to\s+', description)
    
    formatted_parts = []
    for part in parts:
        # Match the Munsell color and the color name
        match = re.match(r'(\d+(\.\d+)?[A-Z]+\s\d+/\d+)\s(.+)', part)
        if match:
            munsell_code = '[' + match.group(1) +']'
            color_name = match.group(3)
            formatted_parts.append(f"{color_name} {munsell_code}")
    
    # Join the formatted parts with 'and' if there are multiple parts
    return ' and '.join(formatted_parts)

# This is for extracting the legend code from [....]
def extract_legend_code(description):
    match = re.search(r'\[(.*?)\]', description)
    return match.group(1) if match else ""

# Remove legend code from the cell data. 
def remove_legend_code(description):
    return re.sub(r'\[.*?\]', '', description).strip()
    

# Function to format and split notes(in table3) into individual points
def split_notes_into_rows(notes):
    # Use regex to split the notes by the pattern (1), (2), (3), ... or *
    split_notes = re.split(r'(\(\d+\)\s*|\*\s*)', notes)
    # Combine the split notes to include the numbering or asterisk with the text, also assemble them into pairs, so the first part of pair is just numbering,
    # like (1), (2) ..... Can be used to match the values in table2 
    formatted_notes = []
    for i in range(1, len(split_notes), 2):
        formatted_notes.append((split_notes[i].strip(), split_notes[i].strip() + split_notes[i+1].strip()))
    return formatted_notes


# Appending table3 data to depth_related_remarks
def process_table_3(table, temp_table2, current_title, depth_related_remarks):
    for row in table[1:]:
        notes = row[0]  
        formatted_notes = split_notes_into_rows(notes)
        
        for note in formatted_notes:
            note_number = note[0]  # The first element of the tuple is the note number
            if len(note_number) == 3:  # Like (1) - (9) 
                note_number = note_number[1:2] # This step is optional, this remove the brackets, so the (1) become 1, to deal with situation like (1,2). This need to be changed if the notes number larger than 10.
            found_match = False
            
            for temp_row in temp_table2[2:]:
                if note_number in temp_row[8]:  # Check if the note number is in the notes column of temp_table2
                    depth_related_remarks.append([current_title, temp_row[0], temp_row[1], note[1]])
                    found_match = True
                    break
            
            if not found_match:
                depth_related_remarks.append([current_title, "", "", note[1]])



def main(doc_path):
    components = list_docx_components(doc_path)
    current_title = None
    # content of the three output files
    location_details = []
    field_geological_descriptions = []
    depth_related_remarks = []
    # use a table counter, count to 3 to determine which table to process
    table_counter = 0
    # temporarily store the table2 for the table3 in the same page to find the depth
    temp_table2 = None

    for component in components:
        if component[0] == 'Paragraph' and len(component[1]) > 1:  #Optional: checking the len(component[1]) is to avoid using some single character ghost as title
            current_title = component[1]
        elif component[0] == 'Table' and current_title:
            table_counter += 1
            table = component[1]
            if table_counter == 1:  # Table 1
                row = [
                    current_title,
                    table[0][1],  # Project No.
                    table[1][1],  # Project
                    table[2][1],  # Client
                    table[0][3],  # Date
                    table[1][3],  # Location
                    table[2][3],  # Logged By
                    table[3][1],  # Excavation Contractor
                    table[3][3],  # Surface Elevation
                    table[4][1],  # Excavation Equipment
                    table[4][3],  # Completion Depth
                    table[5][2],  # Groundwater Depth First Encountered
                    table[6][2],  # Groundwater Depth End of Excavation
                    table[5][4],  # Groundwater Elevation First Encountered
                    table[6][4],  # Groundwater Elevation End of Excavation
                    table[7][1],  # Seasonal High Groundwater Depth
                    table[7][3]   # SHGW Elevation
                ]
                location_details.append(row)
            elif table_counter == 2:  # Table 2
                # temp store or update the temp_table2 for table3 to match the note values
                temp_table2 = table
                for row in table[2:]:
                    munsell_color = reformat_color_description(row[2])
                    legend_code = extract_legend_code(row[5])
                    description_parts = [munsell_color]
                    if row[3] != "-":
                        description_parts.append(row[3])
                    if row[4] != "-":
                        description_parts.append(reformat_color_description(row[4]))
                    description_parts.append(remove_legend_code(row[5]))
                    # Convert row[6] and row[7] to lowercase and append the required strings
                    if row[6] != "-":
                        description_parts.append(", "+ row[6].lower() + " structure")
                    if row[7] != "-":
                        description_parts.append(f"({row[7].lower()})" + " " + "[" + legend_code + "]")
                    description = ' '.join(description_parts).strip()
                    field_geological_descriptions.append([
                        current_title, row[0], row[1], description, legend_code
                    ])
            elif table_counter == 3:  # Table 3
                for row in table[1:]:
                    #depth_related_remarks.append([current_title, "", "", row[0]])
                    process_table_3(table, temp_table2, current_title, depth_related_remarks)
                table_counter = 0  # Reset counter after processing Table 3

    #finally write all data to files.
    write_location_details_to_excel(location_details, "Location_Detail.xlsx")
    write_field_geological_descriptions_to_excel(field_geological_descriptions, "Field_Geological_Descriptions.xlsx")
    write_depth_related_remarks_to_excel(depth_related_remarks, "Depth_Related_Remarks.xlsx")



# For any other files with exact same format, just place this script in the same folder and rename the file name below.
if __name__ == "__main__":
    main("FromDocxTableToExcelFileExample.docx")
